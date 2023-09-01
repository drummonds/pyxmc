from itertools import zip_longest
from openpyxl import load_workbook


def xlsx_same(fn1, fn2):
    same = True
    wb1 = load_workbook(filename = fn1, read_only=True)
    wb2 = load_workbook(filename = fn2, read_only=True)

    if len(wb1.sheetnames) != len(wb2.sheetnames):
        print(f" Difference in number of sheets {len(wb1.sheetnames)} != {len(wb2.sheetnames)}")
        same = False

    for sheet in wb1.sheetnames:
        sheet1 = wb1[sheet]
        try:
            sheet2 = wb2[sheet]
        except KeyError:
            print(f"{fn2} does not have sheet {sheet} which is in {fn1}")
            same = False
            break

        if len(sheet1.rows) != len(sheet2.rows):
            print(f" Difference in number of rows {len(sheet1.rows)} != {len(sheet2.rows)}")
            same = False
        for rownum,row_wb1 in enumerate(sheet1.rows):
            if rownum < sheet2.max_row:
                row_wb2 = sheet2.rows[rownum]

                for colnum, (c1, c2) in enumerate(zip_longest(row_wb1, row_wb2)):
                    if c1 != c2:
                        same = False
                        print(
                            "Row {} Col {} - {} != {}".format(
                                rownum + 1, colnum + 1, c1, c2
                            )
                        )
            else:
                same = False
                print("Row {} missing".format(rownum + 1))
    return same
